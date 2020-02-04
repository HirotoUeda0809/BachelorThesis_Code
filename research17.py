##############################################################
#作成日: 2019/10/22 作成者: Nakahara Nanase
#最終更新日: 2020/1/1 CALC_OJT_2からCALC_OJT_3に切り替え
##############################################################


#ライブラリ読み取りパート
#他ファイル系の読み込み
import RAP_OJT as rp
import CALC_OJT_3 as cl
import DE_OJT as de
import openpyxl as px
path = rp.path()


#パラメーターの読み込みパート
#スタッフ系の情報の読み込み
A = rp.AllMembersList()
ALLNAMELIST = rp.AllMembersNameList()
#日付系の情報の読み取り
MONTH = rp.Month()
N = rp.Days()
SATURDAYS = rp.St_Sn_Hdays("St")
SUNDAYS = rp.St_Sn_Hdays("Sn")
HOLIDAYS = rp.St_Sn_Hdays("H")
#S系(シフト系)
S = rp.Shifts()
SHEETNAME = MONTH
REFRESHARRAY = rp.RefreshArray("{:}月希望.xlsx".format(SHEETNAME))


#解の取得
ANSWER_X = cl.optimal_solution("x")
ANSWER_Y = cl.optimal_solution("y")
ANSWER_G = cl.optimal_solution("g")


#解の出力パート
#workbook(つまりExcelファイル)を生成する．
wb = px.Workbook()
#create_sheet(名前，ページ)で，何ページになんという名前のファイルを作るのか決める．
OutputShiftSheet = wb.create_sheet("OutputShiftSheet", 0)
#罫線を引く
de.draw_line(OutputShiftSheet, A, N,"C")
#色を付ける
de.write_and_color(OutputShiftSheet, A, ALLNAMELIST, N, MONTH, SATURDAYS, HOLIDAYS,"C")
#解の書き込み
de.output_optimal_solution(OutputShiftSheet, A, N, S, ANSWER_X, ANSWER_Y)
#不足をシフト書き込み
de.Lack_S(OutputShiftSheet, A, N, S, ANSWER_G)
#幅調整
de.addjust_width(OutputShiftSheet)


#ファイル保存
wb.save("{:}\\{:}月シフト.xlsx".format(path, MONTH))