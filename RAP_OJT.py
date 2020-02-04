#######################################################################
#更新日:2019/11/1 作成者: Hiroto Ueda(hrt.ueda0809@gmail.com)
#プログラムに使うパラメータや行列の定義などいろいろな関数を集めたファイル
########################################################################

#ライブラリのインポート
import os
import openpyxl as px
import copy
def path():
    path = os.path.expanduser("~\\#SettingsOJT.xlsxを置いた場所")
    return path


#Excelファイルを読み取り
Settings = px.load_workbook("{:}\\SettingsOJT.xlsx".format(path()))
#Settingsの["GeneralSetting"]というシートを読み取る
GeneralSetting = Settings["GeneralSetting"]
#アルバイトスタッフの名前，スキル，契約日数に関する情報を入力するシート
StaffSetting = Settings["StaffSetting"]
#シフトに関する設定
ShiftSetting = Settings["ShiftSetting"]
#研修スタッフに関する設定
TStaffSetting = Settings["TStaffSetting"]


#パラメータの情報を引き出す
#スタッフ系の情報
#数字形式の全員のリスト
def AllMembersList():
    staffnum = GeneralSetting["A20"].value
    all = [i for i in range(1, staffnum + 1)]
    return all
#スタッフ全員の名前のリスト
def AllMembersNameList():
    A = AllMembersList()
    #6列目に名前が保存されているのでcolumn:6
    all_name = [GeneralSetting["F{:}".format(a + 1)].value for a in A]
    return all_name
#各スタッフの契約日数(月)
def StaffContract():
    A = AllMembersList()
    contract = {A[a - 1] : GeneralSetting["G{:}".format(a + 1)].value for a in A}
    return contract
#各スタッフの契約日数(週)
def StaffContract_Week():
    A = AllMembersList()
    contract_week = {A[a - 1] : GeneralSetting["H{:}".format(a + 1)].value for a in A}
    return contract_week

#日付系の情報
def Month():
    Month = GeneralSetting["A5"].value
    return Month
def Days():
    YEAR = GeneralSetting["A3"].value
    MONTH = GeneralSetting["A5"].value
    if (MONTH == 2):
        if (YEAR % 4 == 0 and YEAR % 100 != 0 or YEAR % 400 == 0):
            days = [j for j in range(1, 29 + 1)]
            return days
        else:
            days = [j for j in range(1, 28 + 1)]
            return days
    elif (MONTH == 4 or MONTH == 6 or MONTH == 9 or MONTH == 11):
        days = [j for j in range(1, 30 + 1)]
        return days
    else:
        days = [j for j in range(1, 31 + 1)]
        return days
def St_Sn_Hdays(St_Sn_H):
    DAYS = Days()
    saturdays = []
    sundays = []
    holidays = []
    for k in range(0,5):
        #A7セルにはその月の最初の土曜日の日付が入っている
        satday = GeneralSetting["A7"].value + 7 * k
        sunday = 1 + satday
        if (satday in DAYS):
            saturdays.append(satday)
        if (sunday in DAYS):
            sundays.append(sunday)

    holiday_cell = 9 #祝日の入力はA9セルから
    while (GeneralSetting["A{:}".format(holiday_cell)].value != None):
        holidays += [GeneralSetting["A{:}".format(holiday_cell)].value]
        holiday_cell += 1

    if (St_Sn_H == "St"):
        return saturdays
    elif (St_Sn_H == "Sn"):
        return sundays
    elif (St_Sn_H == "H"):
        return holidays
#S系(シフト系の情報)
def Sinfo():
    shift_info = [GeneralSetting.cell(row = 1, column = j + 2).value for j in range(0, 3)]
    return shift_info
#シフトの集合
def Shifts():
    shifts = []
    for i in range(0, GeneralSetting["A18"].value): #A18セルにはシフト数を入力
        shifts += [GeneralSetting["B{:}".format(i + 2)].value]
    return shifts
#シフトの開始時間を集めた
def ShiftsTime():
    S = Shifts()
    shifts_time = {S[i] : GeneralSetting["D{:}".format(i + 2)].value for i in range(0, len(S))}
    return shifts_time
#シフト名の集合(Trainingも)
def ShiftNameList(Shift_or_Training):
    S = Shifts()
    shifts_name = {S[i]: GeneralSetting["B{:}".format(i + 2)].value for i in range(0, len(S))}
    trainings_name = {S[i]: str(GeneralSetting["B{:}".format(i + 2)].value) + "T" for i in range(0, len(S))}

    if (Shift_or_Training == "S"):
        return shifts_name
    elif (Shift_or_Training == "T"):
        return trainings_name
#平日形態・土日形態などの形態を集めた
def ShiftsTypeName():
    shift_type_name = []
    type_start_num = 3 #F3セルから読み込み開始
    while (ShiftSetting["F{:}".format(type_start_num)].value != None):
        shift_type_name += [ShiftSetting["F{:}".format(type_start_num)].value]
        type_start_num += 1
    return shift_type_name

#勤務希望表の最下段からスケジュール作成日のシフト構成を読み取る
def TodaysShifts(ExcelFileName):
    RefreshSettings = px.load_workbook("{:}\\{:}".format(path(), ExcelFileName))
    HopeShiftMatrix = RefreshSettings["HopeShiftSheet"]
    N = Days()
    A = AllMembersList()
    shift_type = []
    for x in range(0, len(N)):
        shift_type += [HopeShiftMatrix.cell(row = len(A) + 2, column = x + 2).value]
    return shift_type
#当該日のシフトの下限人数と上限人数を決定
def ShiftBound_ts(ExcelFileName, LorU):
    N = Days()
    S = Shifts()
    TS = TodaysShifts(ExcelFileName)
    ST = ShiftsTypeName()
    LowBound = [[0 for s in range(0, len(S))] for j in range(0, len(N))]
    UpBound = [[0 for s in range(0, len(S))] for j in range(0, len(N))]
    for j in range(0, len(N)):
        for s in range(0, len(S)):
            for k in range(0, len(S)):
                if (TS[j] == ST[k]):
                    LowBound[j][s] = ShiftSetting.cell(row = 3 + k, column = 7 + s).value
                    UpBound[j][s] = ShiftSetting.cell(row = 18 + k, column = 7 + s).value
    
    ShiftLowBound = {}
    ShiftUpBound = {}
    for j in N:
        for s in S:
            ShiftLowBound[j,s] = LowBound[N.index(j)][S.index(s)]
            ShiftUpBound[j,s] = UpBound[N.index(j)][S.index(s)]
    if (LorU == "L"):
        return ShiftLowBound
    elif(LorU == "U"):
        return ShiftUpBound
#T系(OJT系の情報)
#研修計画行列
def TPlanArray():
    A = AllMembersList()
    S = Shifts()
    training_plan = [[0 for i in range(0, len(S))] for j in range(0, len(A))]
    for x in range(0, len(A)):
        for y in range(0, len(S)):
            training_plan[x][y] = TStaffSetting.cell(row = x + 3, column = y + 3).value
    training_plan_array = {}
    for a in A:
        for s in S:
            training_plan_array[a,s] = training_plan[A.index(a)][S.index(s)]
    return training_plan_array
#指導スキルに関する行列
def TeachSkillArray():
    A = AllMembersList()
    S = Shifts()
    teach_skill = [[0 for i in range(0, len(S))] for j in range(0, len(A))]
    for x in range(0, len(A)):
        for y in range(0, len(S)):
            teach_skill[x][y] = StaffSetting.cell(row = x + 3, column = y + len(S) + 7).value
    teach_skill_array = {}
    for a in A:
        for s in S:
            teach_skill_array[a,s] = teach_skill[A.index(a)][S.index(s)]
    return teach_skill_array
#各個人の研修回数(論文ではtas-casだがひとまとめに)
def tas():
    A = AllMembersList()
    S = Shifts()
    training_count = [[TStaffSetting.cell(row= a + 3, column = s + len(S) + 6).value for s in range(0, len(S))] for a in range(0, len(A))]
    t_count = {}
    for a in A:
        for s in S:
                t_count[a,s] = training_count[A.index(a)][S.index(s)]
    return t_count
#その他
#休暇とシフト希望行列
def RefreshArray(ExcelFileName):
    RefreshSettings = px.load_workbook("{:}\\{:}".format(path(), ExcelFileName))
    RefreSetting = RefreshSettings["HopeShiftSheet"]
    ALL = AllMembersList()
    N = Days()
    Refresh = [[0 for p in range(0, len(N))] for q in range(0, len(ALL))]
    for s in range(0, len(ALL)):
        for t in range(0, len(N)):
            Refresh[s][t] = RefreSetting.cell(row = s + 2, column = t + 2).value
    RefreshArray = {}
    for i in ALL:
        for j in N:
            RefreshArray[i,j] = Refresh[ALL.index(i)][N.index(j)]
    return RefreshArray
#スタッフのスキルに関する行列
def SkillArray():
    A = AllMembersList()
    S = Shifts()
    skill = [[0 for i in range(0, len(S))] for j in range(0, len(A))]
    for x in range(0, len(A)):
        for y in range(0, len(S)):
            skill[x][y] = StaffSetting.cell(row = x + 3, column = y + 3).value
    skill_array = {}
    for i in A:
        for s in S:
            skill_array[i,s] = skill[A.index(i)][S.index(s)]
    return skill_array
