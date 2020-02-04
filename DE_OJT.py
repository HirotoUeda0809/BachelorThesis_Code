########################################################################
#更新日: 2019/11/1 作成者: Hiroto Ueda(hrt.ueda0809@gmail.com)
#Excelに罫線を引いたり色をつけたりしてくれるプログラム
########################################################################

#ライブラリの読み込み
import openpyxl as px
from openpyxl.styles import (
    PatternFill,
    Border,
    Side,
    )
import RAP_OJT as rp
import random as rd

#シフトの名前
SNL = rp.ShiftNameList("S")
TSNL = rp.ShiftNameList("T")

#色
Yellow = PatternFill(patternType = "solid", fgColor = "ffff00", bgColor = "ffff00")
fillSat = PatternFill(patternType = "solid", fgColor = "87cefa", bgColor = "87cefa")
fillSun = PatternFill(patternType = "solid", fgColor = "ff4500", bgColor = "ff4500")


#Excelファイルに罫線を引くMならMSS，Cならresearch17
def draw_line(ShiftMakeSheet, StaffList, DAYS, MorC):
    border_ln = Border(top = Side(style="thin", color = "000000"),
                    bottom = Side(style="thin", color = "000000"),
                    left = Side(style="thin", color = "000000"),
                    right = Side(style="thin", color = "000000")
                    )
    if (MorC == "M"): #最下段にシフト構成の行を作る
        for row_num in range(1, len(StaffList) + 2 + 1):
            for col_num in range(1, len(DAYS) + 1 + 1):
                ShiftMakeSheet.cell(row = row_num, column = col_num).border = border_ln
    elif (MorC == "C"):
        for row_num in range(1, len(StaffList) + 1 + 1):
            for col_num in range(1, len(DAYS) + 1 + 1):
                ShiftMakeSheet.cell(row = row_num, column = col_num).border = border_ln
#Excelファイルに行と列名を書き込み色を付ける．
#(日曜祝日の×印処理もここ)
def write_and_color(ShiftMakeSheet, StaffList, ANAMELIST, DAYS, MONTH, SATURDAYS, HOLIDAYS, MorC):
    #行にスタッフ名を書き込む
    for i in StaffList:
        ShiftMakeSheet.cell(row = i + 1, column = 1).value = ANAMELIST[i - 1]
    if (MorC == "C"):
        pass
    elif (MorC == "M"):
        ShiftMakeSheet.cell(row = len(StaffList) + 2, column = 1).value = "構成→"
        ShiftMakeSheet.cell(row = len(StaffList) + 2, column = 1).fill = Yellow
    else:
        pass
    
    #列名の書き込みと色付け
    for j in DAYS:
        ShiftMakeSheet.cell(row = 1, column = j + 1).value = "{:}/{:}".format(MONTH, j)
        #土日の色つけ処理
        for sat in SATURDAYS:
            if (j == sat):
                for i in StaffList:
                    ShiftMakeSheet.cell(row = i + 1, column = j + 1).fill = fillSat
            elif (j == sat + 1):
                for i in StaffList:
                    ShiftMakeSheet.cell(row = i + 1, column = j + 1).fill = fillSun
        for holiday in HOLIDAYS:
            if (j == holiday):
                for i in StaffList:
                    ShiftMakeSheet.cell(row = i + 1, column = j + 1).fill = fillSun
#解を受けてどのシフトなのか決定
def output_optimal_solution(ShiftMakeSheet, StaffList, DAYS, SHIFTS, optimal_x, optimal_y):
    for j in DAYS:
        for s in SHIFTS:
            for i in StaffList:
                if (optimal_x[i,j,s].value() == 1):
                    ShiftMakeSheet.cell(row = i + 1, column = j + 1).value = SNL[s]
                if (optimal_y[i,j,s].value() == 1):
                    ShiftMakeSheet.cell(row = i + 1, column = j + 1).value = TSNL[s]
#不足シフト表示
def Lack_S(ShiftMakeSheet, StaffList, DAYS, SHIFTS, optimal_g):
    for j in DAYS:
        k = 0
        for s in SHIFTS:
            if (optimal_g[j,s].value() == 1):
                ShiftMakeSheet.cell(row = len(StaffList) + 2 + k, column = j +1).value = SNL[s]
                k += 1
            else:
                pass
#幅調整
def addjust_width(ShiftMakeSheet):
    for col in ShiftMakeSheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        adjusted_width = (max_length + 2) * 1.5
        ShiftMakeSheet.column_dimensions[column].width = adjusted_width
